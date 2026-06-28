import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

cb = json.load(open("../data/case_bank.json", encoding="utf-8"))
key = json.load(open("../data/key_items.json", encoding="utf-8"))

HEADERS = [
    "증례ID", "분류", "환자(이름/나이/성별)", "주증상", "1순위 진단명", "핵심 진찰 소견",
    "[작성] 환자 평소 말투·성격 한 줄", "[작성] 진찰받을 때 통증 반응 대사 1~2개",
    "[작성] 진단/치료계획 들었을 때 반응 대사 1~2개", "비고",
]
FILL_COLS = [7, 8, 9]  # 1-indexed columns students must fill in

wb = Workbook()
ws = wb.active
ws.title = "입력요청"

header_fill = PatternFill("solid", start_color="4472C4")
header_font = Font(bold=True, color="FFFFFF", name="Arial")
fill_fill = PatternFill("solid", start_color="FFFF00")
wrap = Alignment(wrap_text=True, vertical="top")

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(1, col, h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

for r, c in enumerate(cb["cases"], 2):
    p = c["patient"]
    k = key.get(c["case_id"], {"pe": ""})
    row = [
        c["case_id"], c["category"], f"{p['name']}/{p['age']}/{p['sex']}",
        c["chief_complaint"], c["answer_key"]["top_diagnoses"][0], k["pe"],
        "", "", "", "",
    ]
    for col, val in enumerate(row, 1):
        cell = ws.cell(r, col, val)
        cell.alignment = wrap
        cell.font = Font(name="Arial", size=10)
        if col in FILL_COLS:
            cell.fill = fill_fill

widths = [11, 10, 16, 12, 22, 28, 26, 30, 30, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 34

# 안내 시트
guide = wb.create_sheet("안내")
guide_lines = [
    "■ 작성 안내",
    "노란색 칸(G~I열) 3개만 채워주시면 됩니다. 나머지는 참고용 데이터입니다.",
    "",
    "G열 - 환자 평소 말투·성격: 나이/성별/직업을 반영해서 한 줄로. 예: '교사라 차분하지만 아플 땐 짧게 끊어 말함'",
    "H열 - 진찰받을 때 반응: 해당 진찰 소견(F열)을 실제로 받으면 환자가 보일 통증 반응 대사. 예: '아악! 거기 누르니까 너무 아파요!'",
    "I열 - 진단/치료계획 반응: 의사가 진단명이나 수술/검사 계획을 말했을 때 일반인이 보일 감정 반응 대사. 예: '수술이요? 많이 위험한 거예요?'",
    "",
    "※ I열 작성 시 주의: 환자는 의학 용어를 모르는 일반인입니다. '천공'같은 말 대신 '배가 터졌다고요?' 처럼 일반인이 실제로 할 법한 반응으로 적어주세요.",
    "※ 한 칸에 여러 대사를 적어도 됩니다(케이스 진행 상황에 따라 다르게 쓸 수 있게).",
    "※ 작성 끝나면 이 파일 그대로 다시 전달해주시면 case_bank.json에 반영하겠습니다.",
]
for i, line in enumerate(guide_lines, 1):
    cell = guide.cell(i, 1, line)
    if i == 1:
        cell.font = Font(bold=True, size=13, name="Arial")
    else:
        cell.font = Font(name="Arial", size=11)
guide.column_dimensions["A"].width = 100

wb.save("../student_input/CPX_의대생_입력요청.xlsx")
print("done")
