#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sync_app.py — 단일 원본(data/case_bank.json) → 산출물 동기화

data/case_bank.json 을 유일한 원본으로 보고:
  1) index.html 안의 `const CASE_BANK = {...};` 와 `const KEY_ITEMS = {...};` 블록을 재생성
     (그 외 HTML/JS/프롬프트/스타일은 그대로 보존)
  2) rubric/CPX_채점기준표.xlsx 를 채점기준대로 재생성

KEY_ITEMS 는 각 케이스의 scoring.history_checklist 에서 자동 파생됩니다.

사용법:
    cd scripts
    python3 sync_app.py
또는 프로젝트 루트에서:
    python3 scripts/sync_app.py
"""
import json, os, re, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CASE_BANK_PATH = os.path.join(ROOT, "data", "case_bank.json")
INDEX_PATH = os.path.join(ROOT, "index.html")
XLSX_PATH = os.path.join(ROOT, "rubric", "CPX_채점기준표.xlsx")


def load_source():
    with open(CASE_BANK_PATH, encoding="utf-8") as f:
        data = json.load(f)
    cases = data["cases"] if isinstance(data, dict) and "cases" in data else data
    return cases


def _find_block(src, decl_regex):
    """`<decl> { ... }` 형태에서 균형 잡힌 중괄호 구간의 (start, end) 반환. 문자열 내 중괄호 무시."""
    m = re.search(decl_regex, src)
    if not m:
        raise RuntimeError(f"선언을 찾지 못함: {decl_regex}")
    i = src.index("{", m.end() - 1)
    depth = 0
    instr = False
    esc = False
    q = ""
    for j in range(i, len(src)):
        ch = src[j]
        if instr:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == q:
                instr = False
        else:
            if ch in "\"'":
                instr = True
                q = ch
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return i, j + 1
    raise RuntimeError(f"중괄호 균형이 맞지 않음: {decl_regex}")


def sync_index_html(cases):
    with open(INDEX_PATH, encoding="utf-8") as f:
        html = f.read()

    case_bank_json = json.dumps({"cases": cases}, ensure_ascii=False, separators=(",", ":"))
    key_items = {c["case_id"]: c.get("scoring", {}).get("history_checklist", []) for c in cases}
    key_items_json = json.dumps(key_items, ensure_ascii=False, separators=(",", ":"))

    # CASE_BANK 블록 교체
    a, b = _find_block(html, r"const\s+CASE_BANK\s*=\s*")
    html = html[:a] + case_bank_json + html[b:]
    # KEY_ITEMS 블록 교체 (CASE_BANK 교체 후 인덱스가 바뀌므로 다시 검색)
    a, b = _find_block(html, r"const\s+KEY_ITEMS\s*=\s*")
    html = html[:a] + key_items_json + html[b:]

    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✓ index.html 동기화: CASE_BANK({len(cases)}케이스), KEY_ITEMS 파생 완료")


def sync_rubric_xlsx(cases):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ! openpyxl 미설치 — xlsx 생성을 건너뜀 (pip install openpyxl)")
        return

    wb = Workbook()

    # ── 안내 시트 ──
    ws0 = wb.active
    ws0.title = "안내"
    ws0["A1"] = "CPX 채점기준표 (자동 생성)"
    ws0["A1"].font = Font(size=14, bold=True)
    notes = [
        "",
        "■ 이 파일은 data/case_bank.json 을 원본으로 자동 생성됩니다.",
        "   직접 수정하지 말고, case_bank.json 을 고친 뒤 `python3 scripts/sync_app.py` 를 실행하세요.",
        "",
        "■ 총점 100점 = 병력청취 40 / 신체진찰 20 / PPI(환자소통) 20 / 임상추론 20",
        "■ 병력청취: 아래 항목별 키워드가 학생 발화에 있으면 배점 부여 (항목별 배점 합 = 병력청취 만점)",
        "■ 신체진찰: 시진→청진→타진→촉진 순서 준수 + 양성/음성 소견 확인 여부 평가, 예상 술기 수행 여부",
        "■ 임상추론: 필요검사·치료계획 언급 여부, 소견 근거 기반 진단 추론",
    ]
    for idx, t in enumerate(notes, start=2):
        ws0[f"A{idx}"] = t
    ws0.column_dimensions["A"].width = 90

    # ── 채점기준 시트 ──
    ws = wb.create_sheet("채점기준")
    headers = ["case_id", "환자", "영역", "항목", "키워드", "배점"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="D8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for c in cases:
        cid = c["case_id"]
        p = c.get("patient", {})
        name = f"{p.get('name','')}({p.get('age','')}{p.get('sex','')})"
        sc = c.get("scoring", {})

        # 병력청취 (항목별 배점)
        for it in sc.get("history_checklist", []):
            ws.append([cid, name, "병력청취",
                       it.get("item", ""),
                       ", ".join(it.get("keywords", [])),
                       it.get("score", "")])
        # 신체진찰 예상 술기
        for sk in sc.get("expected_skills", []):
            ws.append([cid, name, "신체진찰", sk, "", ""])
        # 임상추론 - 필요검사 / 치료계획
        if sc.get("required_tests"):
            ws.append([cid, name, "임상추론(검사)", " · ".join(sc["required_tests"]), "", ""])
        if sc.get("required_treatments"):
            ws.append([cid, name, "임상추론(치료)", " · ".join(sc["required_treatments"]), "", ""])

    widths = [12, 16, 14, 40, 32, 6]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"  ✓ rubric/CPX_채점기준표.xlsx 생성: {ws.max_row - 1}행")


def main():
    cases = load_source()
    print(f"원본: data/case_bank.json ({len(cases)} 케이스)")
    sync_index_html(cases)
    sync_rubric_xlsx(cases)
    print("동기화 완료.")


if __name__ == "__main__":
    main()
