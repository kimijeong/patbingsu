import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

cb = json.load(open("../data/case_bank.json", encoding="utf-8"))
cases = {c["case_id"]: c for c in cb["cases"]}

# case-specific key items
KEY = {
 "acute_01": {"hx":"임신 가능성/LMP 확인, 통증 이동 양상(배꼽주변→우하복부), 방사통 여부 문진","pe":"Psoas sign/Obturator sign 시행"},
 "acute_02": {"hx":"기름진 음식/과식과의 연관성, 황달·소변색·가려움 문진","pe":"Murphy's sign 시행"},
 "acute_03": {"hx":"체중감소·지방변·다음/다뇨/다갈 등 췌장기능 관련 문진, 음주력 확인","pe":"DRE 시행"},
 "acute_04": {"hx":"통증 양상(칼로 찢는 듯)과 자세에 따른 변화, 음주력 확인","pe":"DRE 및 복막자극징후 확인"},
 "acute_05": {"hx":"황달/가려움/소변색·대변색 변화 문진","pe":"결막 황달 확인, DRE"},
 "acute_06": {"hx":"임신 가능성/생리 관련 부인과 문진, 통증 이동 양상 확인","pe":"Rovsing/Psoas/Obturator sign 시행"},
 "acute_07": {"hx":"혈뇨/배뇨통 등 비뇨기 증상, 콜릭성 통증 양상 확인","pe":"CVA tenderness 확인"},
 "chronic_01": {"hx":"체중감소 및 식전/공복 시 악화 양상(궤양 시사) 확인","pe":"Epigastric tenderness 확인"},
 "chronic_02": {"hx":"가족력(췌장염/담낭암)과 음주력 확인","pe":"Murphy's sign, CVAT 확인"},
 "chronic_03": {"hx":"체중감소 정도와 경고증상(red flag) 확인","pe":"복부 압통 여부 정밀 확인"},
 "chronic_04": {"hx":"체중감소 및 식후 패턴, 당뇨 조절 상태 확인","pe":"반동압통(rebound tenderness) 여부 명확히 확인"},
 "chronic_05": {"hx":"배변과의 연관성, 스트레스 요인(자녀 문제 등) 확인","pe":"복부 촉진 및 골반 내진"},
 "chronic_06": {"hx":"스트레스/식습관 및 경고증상(체중감소 등) 부재 확인","pe":"Epigastric tenderness 확인"},
 "chronic_07": {"hx":"복용 약물(항생제/소염제 등) 구체적 확인","pe":"DRE 시행(혈변 감별)"},
 "chronic_08": {"hx":"당뇨 조절 상태(HbA1c, 약물 순응도) 확인","pe":"DRE 시행"},
 "acute_08": {"hx":"임신 가능성/LMP 확인, 질출혈 양상, 어지럼/실신감 등 쇼크 징후 문진","pe":"골반 내진 및 자궁목 유동 압통(CMT) 확인"},
 "acute_09": {"hx":"복부 수술력(PSH), 가스 배출/배변 여부 확인","pe":"복부 시진(팽만/수술흉터) 및 청진(금속성 장음) 확인"},
 "acute_10": {"hx":"혈뇨/배뇨통 등 비뇨기 증상과 콜릭성 통증 양상 확인","pe":"CVA tenderness 확인"},
 "acute_11": {"hx":"기름진 음식과의 연관성 및 통증 지속시간(1~5시간) 확인","pe":"Murphy's sign 및 황달 확인"},
 "extra_01": {"hx":"공동 식사자 등 역학적 노출력 확인","pe":"탈수 소견(피부긴장도/구강점막) 평가"},
 "extra_02": {"hx":"성생활력/성매개감염 위험요인 확인, 임신 가능성 배제","pe":"골반 내진(자궁목 유동압통, 부속기 압통) 확인"},
 "extra_03": {"hx":"NSAID/항응고제 복용력, 토혈/흑색변 여부 확인","pe":"직장수지검사(흑색변) 및 기립성 저혈압 등 쇼크 징후 확인"},
 "extra_04": {"hx":"심혈관 위험요인(당뇨/고혈압/흡연/가족력) 및 방사통(어깨/턱) 확인","pe":"심전도 우선 고려 및 발한 등 전신소견 확인"},
}

rows = []  # case_id, 환자, 영역, 항목, 배점, 비고
for cid, c in cases.items():
    name = f"{c['patient']['name']}({c['patient']['age']}{c['patient']['sex']})"
    k = KEY[cid]
    # 병력청취 40
    hx_items = [
        (k["hx"], 8, "케이스 핵심 감별 포인트"),
        ("OPQRST(발병/위치/기간/양상/완화·악화요인) 체계적 문진", 5, ""),
        ("동반 전신증상(발열/체중변화/오한 등) 확인", 4, ""),
        ("동반 소화기/비뇨기 등 계통별 증상 확인", 4, ""),
        ("과거 유사 병력 및 내시경/검진 이력(E) 확인", 4, ""),
        ("복용 약물 확인", 4, ""),
        ("사회력(음주/흡연/직업/식습관) 확인", 4, ""),
        ("가족력 확인" + (" 및 여성력(LMP/임신가능성) 확인" if c["patient"]["sex"]=="여" else ""), 4, ""),
        ("환자의 우려사항/질문 반영 문진", 3, f"환자 질문: {c['patient_question']}"),
    ]
    s = sum(x[1] for x in hx_items)
    if s != 40:
        hx_items[-1] = (hx_items[-1][0], hx_items[-1][1] + (40-s), hx_items[-1][2])
    for item, pt, note in hx_items:
        rows.append([cid, name, "병력청취", item, pt, note])

    # 신체진찰 20
    pe_items = [
        ("General/Vital sign 확인", 5, ""),
        ("복부 시진/청진/타진/촉진 기본 시행", 5, ""),
        (k["pe"], 5, "케이스 핵심 수기"),
        ("필요시 DRE/골반진찰 등 추가 진찰 시행", 5, ""),
    ]
    for item, pt, note in pe_items:
        rows.append([cid, name, "신체진찰", item, pt, note])

    # PPI 20 (실제 CPX 평가지 6항목 기준, 4단계 척도: 아주우수100%/우수70%/보통40%/미흡0%)
    ppi_items = [
        ("내 이야기를 효율적으로 물어보고 잘 들어주었다", 4, "개방형/폐쇄형 질문, 호응, 대답 여유, 확인, 쉬운 용어, 분리 질문, 경청 자세, 면담주제 협상 [비언어 요소 포함→명시 없으면 대화 흐름으로 관대하게 추정]"),
        ("나의 생각과 배경을 효과적으로 알아냈다", 3, "생각/걱정 질문, 기분/정서 표현 격려, 나의 기대 파악, 일상생활 영향 파악, 나의 입장/배경/처지 등에 관심"),
        ("내가 이해하기 쉽게 설명하였다", 4, f"쉬운 용어, 필요한 정보, 내 의견과 선택권 고려, 기억하기 쉽게 설명, 이해 점검 및 질문 기회, 근거 있는 설명 / 환자 질문: {c['patient_question']}"),
        ("나와 좋은 유대관계를 형성하려고 했다", 4, f"편하게 시작, 공감과 지지, 무비판적 수용, 진정성/솔직함, 편안한 분위기, 신뢰, 자신감, 존중 / 특이사항: {c['special_behavior']}"),
        ("면담을 체계적으로 이끌어나갔다", 3, "논리/체계적 순서, 적절한 시간 배분, 주기적 요약/면담 방향 제시, 내 생각에 따라 질문 이어가기"),
        ("신체진찰 태도가 좋았다", 2, "손위생, 사전 설명, 가려주기, 환자안전과 불편함 배려 [손위생/가려주기는 괄호 행동 표기 시만 인정, 그 외엔 전반적 태도로 관대하게 추정]"),
    ]
    for item, pt, note in ppi_items:
        rows.append([cid, name, "PPI(환자-의사관계)", f"{item} (척도: 아주우수=100%/우수=70%/보통=40%/미흡=0%)", pt, note])

    # 임상추론 20
    ak = c["answer_key"]
    dx = ak["top_diagnoses"]
    tests = ak["required_tests"]
    tx = ak["required_treatments"] or ak["expected_skills"]
    reasoning_items = [
        (f"1순위 추정진단 언급: {dx[0] if dx else '-'}", 6, ""),
        (f"2~3순위 감별진단 언급: {', '.join(dx[1:]) if len(dx)>1 else '-'}", 3, ""),
        (f"핵심 검사 제안: {', '.join(tests) if tests else '-'}", 6, ""),
        (f"치료계획/예상술기 언급: {', '.join(tx) if tx else '-'}", 5, ""),
    ]
    for item, pt, note in reasoning_items:
        rows.append([cid, name, "임상추론(Dx/검사/치료)", item, pt, note])

wb = Workbook()
FONT = "Arial"

# Sheet 1: 안내
ws0 = wb.active
ws0.title = "안내"
ws0.column_dimensions['A'].width = 100
intro = [
 "CPX 복통 증례 채점 기준 (검토용)",
 "",
 "■ 채점 영역 및 배점 (100점 만점)",
 "  1) 병력청취(History Taking) - 40점",
 "  2) 신체진찰(Physical Exam) - 20점",
 "  3) 환자-의사관계(PPI: 공감/설명/태도) - 20점",
 "  4) 임상추론(감별진단·검사계획·치료계획) - 20점",
 "",
 "■ 사용 방법",
 "  - '세부체크리스트' 시트: 증례별 항목과 배점이 나열되어 있습니다. 배점(E열) 숫자만 수정하면 됩니다.",
 "  - '증례별총점' 시트: 세부체크리스트의 배점을 영역별로 합산하여 100점 만점 점수를 계산합니다 (수식 자동 반영).",
 "  - 항목 문구나 배점은 자유롭게 수정/추가/삭제 가능합니다. 단, 영역별 합계가 의도한 가중치(40/20/20/20)와 다르면 '증례별총점' 시트에서 바로 확인할 수 있습니다.",
 "",
 "■ 참고",
 "  - 각 증례의 '케이스 핵심 감별 포인트', '케이스 핵심 수기' 항목은 교재의 [시험관련 코멘트]를 반영해 다른 항목보다 배점을 높게 잡았습니다.",
 "  - LLM 연동(환자 시뮬레이터/자동 채점)은 이 채점표가 확정된 후 진행합니다.",
 "",
 "■ LLM 채점 가이드 (채팅 기반 CPX의 한계 반영)",
 "  - 이 CPX는 텍스트 채팅으로 진행되므로, 실제 손위생/표정/말투/시진 등 비언어적 행동을 LLM이 직접 관찰할 수 없습니다.",
 "  - PPI 6항목 중 '경청 자세/대답 여유'(1번), '손위생/가려주기/환자안전배려'(6번, 신체진찰 태도)는 학생이 괄호 행동 표기(예: \"(손소독제를 사용한다)\", \"(검진포로 가려준다)\")로 명시하면 그 텍스트를 근거로 채점합니다.",
 "  - 명시적 언급이 없는 경우, LLM은 대화 전체의 흐름과 어투(질문이 급박하거나 무례하지 않은지, 환자를 배려하는 멘트가 있는지 등)를 바탕으로 관대하게 추정하여 점수를 부여합니다(엄격한 미흡 처리 대신 보통 수준 이상을 기본값으로 적용).",
 "  - 단, 촉진/타진처럼 구역(좌/우, 상/중/하)을 명시해야 하는 신체진찰 항목은 추정 없이 텍스트에 구역이 실제로 언급되었는지로 엄격히 채점합니다.",
]
for i, line in enumerate(intro, start=1):
    cell = ws0.cell(row=i, column=1, value=line)
    cell.font = Font(name=FONT, bold=(i==1 or line.startswith("■")), size=14 if i==1 else 11)

# Sheet 2: 세부체크리스트
ws1 = wb.create_sheet("세부체크리스트")
headers = ["증례ID", "환자", "영역", "체크리스트 항목", "배점", "비고"]
ws1.append(headers)
for col in range(1,7):
    c = ws1.cell(row=1, column=col)
    c.font = Font(name=FONT, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="4472C4")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

domain_fill = {
    "병력청취": "DDEBF7",
    "신체진찰": "E2EFDA",
    "PPI(환자-의사관계)": "FFF2CC",
    "임상추론(Dx/검사/치료)": "FCE4D6",
}
for r in rows:
    ws1.append(r)
for i in range(2, len(rows)+2):
    dom = ws1.cell(row=i, column=3).value
    for col in range(1,7):
        cell = ws1.cell(row=i, column=col)
        cell.font = Font(name=FONT, size=10)
        cell.fill = PatternFill("solid", start_color=domain_fill.get(dom, "FFFFFF"))
        if col in (4,6):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if col == 5:
            cell.alignment = Alignment(horizontal="center")

widths = {"A":12,"B":16,"C":20,"D":55,"E":8,"F":40}
for k,v in widths.items():
    ws1.column_dimensions[k].width = v
ws1.freeze_panes = "A2"

# Sheet 3: 증례별총점
ws2 = wb.create_sheet("증례별총점")
ws2.append(["증례ID","환자","병력청취(40)","신체진찰(20)","PPI(20)","임상추론(20)","총점(100)"])
for col in range(1,8):
    c = ws2.cell(row=1, column=col)
    c.font = Font(name=FONT, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="4472C4")
    c.alignment = Alignment(horizontal="center")

case_ids = list(cases.keys())
for ridx, cid in enumerate(case_ids, start=2):
    name = f"{cases[cid]['patient']['name']}({cases[cid]['patient']['age']}{cases[cid]['patient']['sex']})"
    ws2.cell(row=ridx, column=1, value=cid)
    ws2.cell(row=ridx, column=2, value=name)
    domains = ["병력청취","신체진찰","PPI(환자-의사관계)","임상추론(Dx/검사/치료)"]
    for dcol, dom in enumerate(domains, start=3):
        formula = f'=SUMIFS(세부체크리스트!$E$2:$E${len(rows)+1},세부체크리스트!$A$2:$A${len(rows)+1},$A{ridx},세부체크리스트!$C$2:$C${len(rows)+1},"{dom}")'
        ws2.cell(row=ridx, column=dcol, value=formula)
    ws2.cell(row=ridx, column=7, value=f"=SUM(C{ridx}:F{ridx})")
    for col in range(1,8):
        cell = ws2.cell(row=ridx, column=col)
        cell.font = Font(name=FONT, size=11)
        if col >= 3:
            cell.alignment = Alignment(horizontal="center")

last = len(case_ids)+1
ws2.append(["평균","", f"=AVERAGE(C2:C{last})", f"=AVERAGE(D2:D{last})", f"=AVERAGE(E2:E{last})", f"=AVERAGE(F2:F{last})", f"=AVERAGE(G2:G{last})"])
for col in range(1,8):
    cell = ws2.cell(row=last+1, column=col)
    cell.font = Font(name=FONT, bold=True)
    if col>=3: cell.alignment = Alignment(horizontal="center")

widths2 = {"A":12,"B":16,"C":14,"D":14,"E":10,"F":14,"G":10}
for k,v in widths2.items():
    ws2.column_dimensions[k].width = v

wb.save("../rubric/CPX_채점기준표.xlsx")
print("rows:", len(rows))
